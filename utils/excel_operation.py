import openpyxl

class Excelopration:
    @staticmethod
    def read_excel(path):
        excel = openpyxl.load_workbook(path)
        sheet = excel['Sheet1']
        max_row = sheet.max_row
        # print('行数:',max_row)
        max_column = sheet.max_column
        # print('列数：',max_column)
        data_list=[]

        for i in range(2,max_row+1):
            row_list=[]
            for j in range(1,max_column+1):
                a = sheet.cell(row = i,column=j).value
                # print(a)
                row_list.append(a)
            data_list.append(row_list)
        return data_list




if __name__ == '__main__':
    print(Excelopration.read_excel('../data/aa.xlsx'))
    print(Excelopration.read_excel('../data/wx.xlsx'))


